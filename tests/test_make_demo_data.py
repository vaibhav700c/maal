import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

import make_demo_data as demo  # noqa: E402

INPUT_CSV = Path(__file__).resolve().parents[1] / "input" / "Unihack_ Sample Dataset - Input.csv"


def test_demo_artifacts_match_web_console_expectations(tmp_path):
    out_dir = tmp_path / "output"
    rows = demo.write_demo_artifacts(INPUT_CSV, out_dir, limit=40)
    assert len(rows) == 40

    csv_path = out_dir / "result.csv"
    xlsx_path = out_dir / "result.xlsx"
    sidecar_path = out_dir / "sidecar.jsonl"
    state_path = out_dir / "state.jsonl"
    assert csv_path.exists() and xlsx_path.exists() and sidecar_path.exists() and state_path.exists()

    # web/lib/artifacts.ts: result.csv must parse with a MANUFACTURER_PART_NUMBER column
    with open(csv_path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        csv_rows = list(reader)
    assert len(csv_rows) == 40
    assert len(reader.fieldnames) == 252
    mpns_csv = {r["MANUFACTURER_PART_NUMBER"] for r in csv_rows}

    # web/lib/artifacts.ts: sidecar.jsonl lines must be JSON with the fields
    # SidecarRecord reads (mfg_part_num, fields, physics, flags, triage_score)
    sidecar_lines = sidecar_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(sidecar_lines) == 40
    mpns_sidecar = set()
    for line in sidecar_lines:
        rec = json.loads(line)
        for key in ("mfg_part_num", "fields", "physics", "flags", "triage_score"):
            assert key in rec
        mpns_sidecar.add(rec["mfg_part_num"])
        # honesty: nothing here was verified against a manufacturer source
        for field in rec["fields"].values():
            assert field["verdict"] == "UNVERIFIED"
            assert field["source_url"] is None
            assert field["tier"] == 0.0
        assert rec["retrieval"] is None
        assert rec["classification"] is None
        assert "DEMO_INPUT_ONLY" in rec["flags"]

    assert mpns_csv == mpns_sidecar

    # web/app/about/page.tsx reads output/state.jsonl the same way run_batch's
    # Checkpoint writes it: {"mfg_part_num": ..., "row_result": {...}}
    state_lines = state_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(state_lines) == 40
    for line in state_lines:
        rec = json.loads(line)
        assert "mfg_part_num" in rec
        assert "extraction" in rec["row_result"]

    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws.max_row == 41  # header + 40 rows
    assert ws.max_column == 252
